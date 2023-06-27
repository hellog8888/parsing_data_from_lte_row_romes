import psycopg2.extras
import datetime
import glob
import openpyxl
import warnings

warnings.simplefilter("ignore")

dict_for_operator = \
    {
        'Общество с ограниченной ответственностью «Скартел»': 'Скартел',
        'Общество с ограниченной ответственностью \"Скартел\"': 'Скартел',

        'Общество с ограниченной ответственностью \"Т2 Мобайл\"': 'Т2 Мобайл',
        'Общество с ограниченной ответственностью «Т2 Мобайл»': 'Т2 Мобайл',

        'Публичное акционерное общество «Мобильные ТелеСистемы»': 'МТС',
        'Публичное акционерное общество \"Мобильные ТелеСистемы\"': 'МТС',

        'Публичное акционерное общество \"МегаФон\"': 'МегаФон',
        'Публичное акционерное общество «МегаФон»': 'МегаФон',

        'Публичное акционерное общество \"Ростелеком\"': 'Ростелеком',
        'Публичное акционерное общество «Ростелеком»': 'Ростелеком',
        'Публичное акционерное общество междугородной и международной электрической связи \"Ростелеком\"': 'Ростелеком',

        'Публичное акционерное общество «Вымпел-Коммуникации»': 'ВымпелКом',
        'Публичное акционерное общество \"Вымпел-Коммуникации\"': 'ВымпелКом'
    }

dict_ETC = \
    {
        '18.1.1.3.': 'GSM',
        '18.1.1.8.': 'GSM',
        '18.1.1.5.': 'UMTS',
        '18.1.1.6.': 'UMTS',
        '18.7.1.': 'LTE',
        '18.7.4.': 'LTE',
        '18.7.5.': '5G NR',
        '19.2.': 'РРС'
    }


def measure_time(func):
    def wrapper(*args, **kwargs):
        start_time = datetime.datetime.now()
        result = func(*args, **kwargs)
        end_time = datetime.datetime.now()
        elapsed_time = end_time - start_time
        print(f"Execution time: {elapsed_time}")
        return result

    return wrapper


@measure_time
def convert_to_postgres(file_open):
    hostname = 'localhost'
    database = 'eirs'
    username = 'postgres'
    pwd = '1234'
    port_id = 5432
    conn = None

    try:
        with psycopg2.connect(
                host=hostname,
                dbname=database,
                user=username,
                password=pwd,
                port=port_id
        ) as conn:

            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:

                # удалить таблицу (не считать ошибкой)
                cur.execute('DROP TABLE IF EXISTS cellular')

                create_script = """ CREATE TABLE IF NOT EXISTS cellular (
                                        РЭС                       varchar(41),
                                        Адрес                     varchar(230),
                                        ТИП_РЭС                   varchar(5),
                                        Владелец                  varchar(11),
                                        Широта                    varchar(9),
                                        Долгота                   varchar(9),
                                        Частоты                   varchar(756),
                                        Дополнительные_параметры  varchar(490),
                                        Классы_излучения          varchar(53),
                                        Серия_Номер_РЗ_СоР        varchar(13))
                                """

                cur.execute(create_script)

                file_to_read = openpyxl.load_workbook(file_open, data_only=True)
                sheet = file_to_read['SQL Results']

                for row in range(2, sheet.max_row + 1):
                    data = []

                    for col in range(1, sheet.max_column + 1):
                        value = sheet.cell(row, col).value
                        data.append(value)

                    cur.execute(
                         "INSERT INTO cellular (РЭС, Адрес, ТИП_РЭС, Владелец, Широта, Долгота, Частоты, Дополнительные_параметры, Классы_излучения, Серия_Номер_РЗ_СоР) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                         (str(data[1]), str(data[2]), str(dict_ETC[data[3]]), str(dict_for_operator[data[6]]),
                          str(data[7]), str(data[8]), str(data[10]), str(data[11]), str(data[17]), f'{data[18]} {data[19]}'))

                    data.clear()

                conn.commit()

    except Exception as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()


file_xlxl_1 = glob.glob('source_folder\*.xlsx')
convert_to_postgres(file_xlxl_1[0])
